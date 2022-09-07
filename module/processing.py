import os, csv
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import NamedStyle
from openpyxl.utils.cell import get_column_letter
from module.timeinfo import TimeInfo



class CsvDataSheet:
    def __init__(self, type, name, path) -> None:
        self.type:str = type
        self.name:str = name
        self.path:str = path
        self.dirPath:str = path[:-(len(name)+1)]


class TimeInfoNotMatched(Exception):
    def __str__(self) -> str:
        return "[오류] 실험 시작시간이 다른 파일이거나, 설정한 시간 이후에 기록된 데이터 밖에 없는 것 같습니다."



def filter_csv_from_dir(dirPath, file_type) -> list:
    processed_file_objects = []

    # 경로안에 있는 유효한 파일들 걸러내기
    fileList = os.listdir(dirPath)
    csvList = [e for e in fileList if ".csv" in e[-4:]]

    # 유효한 파일의 이름 또는 선택한 선택지로 이 데이터의 종류 결정하기
    for file in csvList:
        if "ID" in file:
            processed_file_objects.append(CsvDataSheet("dummy", file, f"{dirPath}/{file}"))
        elif "Total" in file:
            processed_file_objects.append(CsvDataSheet("total", file, f"{dirPath}/{file}"))
        elif 'CO2' in file:
            processed_file_objects.append(CsvDataSheet("CO2", file, f"{dirPath}/{file}"))
            continue # 혹시 O2까지 내려갈까봐 적어둠
        elif 'He' in file:
            processed_file_objects.append(CsvDataSheet("He", file, f"{dirPath}/{file}"))
        elif 'O2' in file:
            processed_file_objects.append(CsvDataSheet("O2", file, f"{dirPath}/{file}"))
        else:
            processed_file_objects.append(CsvDataSheet(file_type, file, f"{dirPath}/{file}")) # 나머지는 본래 입력한 파일 속성으로 할당

    return processed_file_objects



def preprocess_csv_to_df(csvDataSheet:CsvDataSheet, time:TimeInfo) -> pd.DataFrame: # CSV파일 가공하는 함수
    if csvDataSheet.type == "dummy":
        df = pd.read_csv(csvDataSheet.path, encoding="CP949")
        df = df.drop([df.columns[5], df.columns[6], df.columns[7]], axis='columns') # 온도, 습도, 대기압 삭제

        # 시작 시간 추출
        if time.isExist() == False:
            dateInCell = list(map(int, (df.iloc[1,0].split("-"))))
            time.set_time(datetime(dateInCell[0], dateInCell[1], dateInCell[2], time.timedatalist[0], time.timedatalist[1], 0))

        # 측정시간 Datetime으로 변환
        df[df.columns[0]] = pd.to_datetime(df[df.columns[0]], format="%Y-%m-%d-%H-%M-%S")
        
        try:
            # CASE1 시작은 -1초씩 하면서 발견한 값을 시작 0초로 만들고, 끝은 +1초씩 하면서 발견한 값을 끝 0초로 만든다.
            startIndexNum, endIndexNum = [None, None]
            for s in range(11): # 약 10초 범위로 진행
                startIndex = df[df["측정시간"] == time.get_time() - timedelta(seconds=s)] # 시작시간 인덱스 가져오기
                if len(startIndex) > 0:
                    startIndexNum = startIndex.index[0]
                    if s > 0:
                        df.iloc[startIndexNum, df.columns.get_loc('측정시간')] = time.get_time()
                    break
            for s in range(11): # 약 10초 범위로 진행
                endIndex = df[df["측정시간"] == time.get_end_time() + timedelta(seconds=s)] # 시작시간 인덱스 가져오기
                if len(endIndex) > 0:
                    endIndexNum = endIndex.index[0]
                    if s > 0:
                        df.iloc[endIndexNum, df.columns.get_loc('측정시간')] = time.get_end_time()
                    break
            if startIndexNum == None:
                raise TimeInfoNotMatched
            if endIndexNum == None:
                print(f'\r[경고] 실험 종료 시간을 찾을 수 없습니다. {time.get_time().strftime("%H:%M")}에 시작한 후 {time.get_duration()}초 후인 {time.get_end_time().strftime("%H:%M")}에 값이 없습니다. 이 파일의 마지막 시간 정보는 {df.iloc[df.shape[0]-1, 0].strftime("%H:%M:%S")}입니다.  마지막 시간 정보까지만 데이터를 가공합니다.')
                df = df.iloc[startIndexNum:]
                return df
        except:
            raise TimeInfoNotMatched

        df = df.iloc[startIndexNum:endIndexNum+1]
        return df

    elif csvDataSheet.type == "total":
        df = pd.read_csv(csvDataSheet.path, encoding="CP949")
        dropArray = []
        for i in [5,6,7, 12,13,14, 19,20,21, 26,27,28, 33,34,35, 40,41,42, 47,48,49, 54,55,56, 61,62,63, 68,69,70]:
            dropArray.append(df.columns[i])
        df = df.drop(dropArray, axis='columns') # 온도, 습도, 대기압 전체 삭제
        
        # 시작 시간 추출
        if time.isExist() == False:
            dateInCell = list(map(int, (df.iloc[1,0].split("-"))))
            time.set_time(datetime(dateInCell[0], dateInCell[1], dateInCell[2], time.timedatalist[0], time.timedatalist[1], 0))

        # 측정시간 Datetime으로 변환
        df[df.columns[0]] = pd.to_datetime(df[df.columns[0]], format="%Y-%m-%d-%H-%M-%S")
        
        try:
            # CASE1 시작은 -1초씩 하면서 발견한 값을 시작 0초로 만들고, 끝은 +1초씩 하면서 발견한 값을 끝 0초로 만든다.
            startIndexNum, endIndexNum = [None, None]
            for s in range(11): # 약 10초 범위로 진행
                startIndex = df[df["측정시간"] == time.get_time() - timedelta(seconds=s)] # 시작시간 인덱스 가져오기
                if len(startIndex) > 0:
                    startIndexNum = startIndex.index[0]
                    if s > 0:
                        df.iloc[startIndexNum, df.columns.get_loc('측정시간')] = time.get_time()
                    break
            for s in range(11): # 약 10초 범위로 진행
                endIndex = df[df["측정시간"] == time.get_end_time() + timedelta(seconds=s)] # 시작시간 인덱스 가져오기
                if len(endIndex) > 0:
                    endIndexNum = endIndex.index[0]
                    if s > 0:
                        df.iloc[endIndexNum, df.columns.get_loc('측정시간')] = time.get_end_time()
                    break
            if startIndexNum == None:
                raise TimeInfoNotMatched
            if endIndexNum == None:
                print(f'\r[경고] 실험 종료 시간을 찾을 수 없습니다. {time.get_time().strftime("%H:%M")}에 시작한 후 {time.get_duration()}초 후인 {time.get_end_time().strftime("%H:%M")}에 값이 없습니다. 이 파일의 마지막 시간 정보는 {df.iloc[df.shape[0]-1, 0].strftime("%H:%M:%S")}입니다. 마지막 시간 정보까지만 데이터를 가공합니다.')
                df = df.iloc[startIndexNum:]
                return df
        except:
            raise TimeInfoNotMatched
        df = df.iloc[startIndexNum:endIndexNum+1]
        return df


    elif csvDataSheet.type == "CO2" or csvDataSheet.type == "He" or csvDataSheet.type == "O2":
        # csv 전처리 후 numpy Array로 변환하여 DataFrame 생성
        rows = []
        col = {"exist": False, "array": []}
        with open(csvDataSheet.path, 'r') as f:
            rdr = csv.reader(f)
            for i in rdr:
                if col["exist"] == False:
                    del i[1]
                    col["array"] = i
                    col["array"] = list(filter(len, col["array"]))
                    col["exist"] = True
                    continue
                if len(i) > 6:
                    for _ in range(len(i) - 6): # Column이 6개이므로 그 이상은 버리기
                        i.pop()
                i[2] = f"{i[1]} {i[2]}" # 날짜 시간 합쳐주기
                del i[1]
                rows.append(i)
        ar = np.array(rows)
        df = pd.DataFrame(ar, columns=col["array"])
        df = df.astype({"Temp.":float, "Humidity":float})
        df = df.astype({csvDataSheet.type:float})
        # 22.09.06 수정요청에 의해서 아래 내용 주석처리 # FIXME: 여기서 처리하고 아래에서 다시 바꾼다던지 할까?
        # type_kor_name = {'CO2':'이산화탄소(ppm)', 'He':'헬륨(%)', 'O2':'산소(%)'}
        # df.rename(columns={csvDataSheet.type:type_kor_name[csvDataSheet.type]})

        # 시작 시간 추출
        if time.isExist() == False:
            dateInCell = list(map(int, (df.iloc[0,1].split(" "))[0].split("-")))
            time.set_time(datetime(dateInCell[2], dateInCell[0], dateInCell[1], time.timedatalist[0], time.timedatalist[1], 0))

        df = df.drop([df.columns[0]], axis='columns') # 인덱스 삭제
        df[df.columns[0]] = pd.to_datetime(df[df.columns[0]], format="%m-%d-%Y %H:%M:%S") # 측정시간 Datetime으로 변환
        
        # 시작시간 인덱스 가져오기
        try:
            startIndexNum = df[df["TIME"] == time.get_time()].index[0] 
        except:
            raise TimeInfoNotMatched

        indexCount = 0 # 범위에 포함되는 인덱스 잘라내기 위한 변수
        for i in range(startIndexNum, df.shape[0]):
            diff = df.iloc[i,0] - df.iloc[startIndexNum,0]
            if diff.seconds > time.get_duration():
                # print(f"{time.get_duration()}초를 넘어 정지된 시점 : {df.iloc[i,0]}") # FOR DEBUG
                break
            # print(f"[{startIndexNum + indexCount}] {df.iloc[i,0].strftime('%H:%M:%S')}는 시작시간 {df.iloc[startIndexNum,0].strftime('%H:%M:%S')}과의 차가 {diff}초 입니다.") # FOR DEBUG
            indexCount += 1
        df = df.iloc[startIndexNum:startIndexNum + indexCount]
        return df


def calculate_df(df: pd.DataFrame) -> dict:
    df = df.reset_index(drop=True)

    valueArr = df['CO2']
    prev = None
    detected_idx = None
    for i, v in enumerate(valueArr):
        if prev == None:
            prev = valueArr[0]
            continue
        if v >= prev + 10:
            detected_idx = i
            break
        prev = v

    detected_at: datetime = df.iloc[detected_idx, 0]
    detected_period_timedelta: timedelta = detected_at - df.iloc[0, 0]
    detected_period: int = detected_period_timedelta.seconds

    max_value_index = df[df['CO2'] >= 10000].index[0]
    max_value_reached_at: datetime = df.iloc[max_value_index, 0]
    max_value_reached_period_timedelta: timedelta = max_value_reached_at - df.iloc[0, 0]
    max_value_reached_period: int = max_value_reached_period_timedelta.seconds

    from_detected_to_max_period_timedelta: timedelta = max_value_reached_at - detected_at
    from_detected_to_max_period: int = from_detected_to_max_period_timedelta.seconds

    return {
        'detected_at': detected_at,
        'detected_period': detected_period,
        'max_value_reached_at': max_value_reached_at,
        'max_value_reached_period': max_value_reached_period,
        'from_detected_to_max_period': from_detected_to_max_period
    }



def dataframe_to_excel(dataframe:pd.DataFrame):
    try:
        wb = Workbook()
        ws = wb.active
        for i in dataframe_to_rows(dataframe, index=False, header=True):
            ws.append(i)
        return wb

    except:
        print("오류!!! None타입을 받았거나 그 외의 오류 발생")



def formular_process(wb:Workbook):
    try:
        # 열추가, 수식추가
        ws = wb.active
        ws.insert_cols(2, 2)
        ws["C1"] = "경과시간(s)"
        for i in range(2, ws.max_row + 1):
            ws[f"C{i}"] = f"=ROUND((A{i}-$A$2)*24*60*60,0)"
        return wb
    except:
        pass

# FIXME: 여기까지 일단 분석함

def firebase_process(wb: Workbook, csvDataSheet: CsvDataSheet, __type: str) -> dict:
    dataDict = {}
    ws = wb.active

    # 경과시간(s)를 int 타입으로 변경
    for i in range(2, ws.max_row + 1):
        diff = ws[f"A{i}"].value - ws["A2"].value
        ws[f"C{i}"] = diff.seconds

    # 모든 날짜를 str 타입으로 변경
    for i in range(2, ws.max_row + 1):
        ws[f"A{i}"] = ws[f"A{i}"].value.strftime('%H:%M:%S')

    if __type == 'total':
        # pandas에서 중복된 column에 붙인 숫자 제거
        alignNum = 4
        for _ in range(10):
            ws[f'{get_column_letter(alignNum)}1'] = 'ID'
            ws[f'{get_column_letter(alignNum+1)}1'] = '산소(%)'
            ws[f'{get_column_letter(alignNum+2)}1'] = '이산화탄소(ppm)'
            ws[f'{get_column_letter(alignNum+3)}1'] = '헬륨(%)'
            alignNum += 4

        for id in range(10):
            all = ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=0, max_col=ws.max_column, values_only=True) # for 문 바깥에 있었는데 enumerate 후 값이 사라져 다시 선언하기로 하였다.
            dataDict[f'DUMMY_ID{id+1}'] = {}
            for i, row in enumerate(all):
                # 0 - 날짜, 2 - 경과시간 / n - ID, n+1 - O2, n+2 - CO2, n+3 - He
                dataDict[f'DUMMY_ID{id+1}'][i] = [row[0],row[2],row[4 + 4*id],row[5 + 4*id],row[6 + 4*id]]
        return dataDict

    # elif __type == 'CO2': #TODO: 추후 데이터시트를 보고 추가
    # elif __type == 'CO2' or __type == 'He':
        # all = ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=0, max_col=ws.max_column, values_only=True)
        # __si = csvDataSheet.name.rfind('_') + 1 # FIXME: 사용불가능, 이름에 규칙이 없기 때문
        # __ei = csvDataSheet.name.find('.csv')
        # dataDict[f'{__type}_{csvDataSheet.name[__si:__ei]}'] = {}
        # for i, row in enumerate(all):
        #     # 0 - 날짜, 2 - 경과시간, 5 - 타겟가스
        #     dataDict[f'{__type}_{csvDataSheet.name[__si:__ei]}'][i] = [row[0],row[2],row[5]]
        # return dataDict



def expression_process(wb:Workbook, csvDataSheet:CsvDataSheet):
    ws = wb.active

    if csvDataSheet.type == "dummy":
        dateStyle = NamedStyle(name="datetime", number_format="[$-x-systime]h:mm:ss AM/PM")
        for i in range(2, ws.max_row + 1):
            ws[f"A{i}"].style = dateStyle
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 4
        ws.column_dimensions["L"].width = 5
        ws.column_dimensions["W"].width = 1
        return wb
    
    elif csvDataSheet.type == "total":
        dateStyle = NamedStyle(name="datetime", number_format="[$-x-systime]h:mm:ss AM/PM")
        for i in range(2, ws.max_row + 1):
            ws[f"A{i}"].style = dateStyle
        ws.column_dimensions["A"].width = 20

        # pandas에서 중복된 column에 붙인 숫자 제거
        alignNum = 4
        for i in range(1, 11):
            ws[f'{get_column_letter(alignNum)}1'] = 'ID'
            ws[f'{get_column_letter(alignNum+1)}1'] = f'ID{i} 산소(%)'
            ws[f'{get_column_letter(alignNum+2)}1'] = f'ID{i} 이산화탄소(ppm)'
            ws[f'{get_column_letter(alignNum+3)}1'] = f'ID{i} 헬륨(%)'
            alignNum += 4
        return wb
    
    elif csvDataSheet.type == "CO2" or csvDataSheet.type == "He":
        dateStyle = NamedStyle(name="datetime", number_format="[$-x-systime]h:mm:ss AM/PM")
        for i in range(2, ws.max_row + 1):
            ws[f"A{i}"].style = dateStyle
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["C"].width = 10
        return wb


def insert_calculated_values(wb: Workbook, cv: dict, isMeter=False, csvDataSheet:CsvDataSheet=None):
    ws = wb.active
    
    if not isMeter:
        ws['M2'] = '계측기 반응시간'
        ws['N2'] = cv['detected_at'].strftime('%H:%M:%S')
        ws['M3'] = '반응까지 경과시간(초)'
        ws['N3'] = cv['detected_period']
        ws['M4'] = '최대 ppm 도달시간'
        ws['N4'] = cv['max_value_reached_at'].strftime('%H:%M:%S')
        ws['M5'] = '최댓값까지 경과시간(초)'
        ws['N5'] = cv['max_value_reached_period']
        ws['M6'] = '반응 후 최댓값 도달시간(초)'
        ws['N6'] = cv['from_detected_to_max_period']
        ws.column_dimensions['M'].width = 25
        return wb
    else:
        row_count = 4
        while True:
            if ws[f'B{row_count}'] == '' or ws[f'B{row_count}'] == None:
                ws[f'B{row_count}'] = 'ID'
                ws[f'B{row_count+1}'] = '계측기 반응시간'
                ws[f'B{row_count+2}'] = '반응까지 경과시간(초)'
                ws[f'B{row_count+3}'] = '최대 ppm 도달시간'
                ws[f'B{row_count+4}'] = '최댓값까지 경과시간(초)'
                ws[f'B{row_count+5}'] = '반응 후 최댓값 도달시간(초)'
                ws[f'C{row_count}'] = csvDataSheet.name[-1]
                ws[f'C{row_count+1}'] = cv['detected_at'].strftime('%H:%M:%S')
                ws[f'C{row_count+2}'] = cv['detected_period']
                ws[f'C{row_count+3}'] = cv['max_value_reached_at'].strftime('%H:%M:%S')
                ws[f'C{row_count+4}'] = cv['max_value_reached_period']
                ws[f'C{row_count+5}'] = cv['from_detected_to_max_period']
                break
            else:
                row_count += 22
        return wb
        

def chart_process(wb:Workbook, csvDataSheet:CsvDataSheet):
        if csvDataSheet.type == "dummy":
            try:
                # 차트추가
                ws = wb.active
                valueOxygen = Reference(ws, min_row=1, max_row=ws.max_row, min_col=5, max_col=5)
                valueCarbonDioxide = Reference(ws, min_row=1, max_row=ws.max_row, min_col=6, max_col=6)
                valueHelium = Reference(ws, min_row=1, max_row=ws.max_row, min_col=7, max_col=7)
                valueTime = Reference(ws, min_row=2, max_row=ws.max_row, min_col=3, max_col=3)

                oxygenChart = LineChart()
                oxygenChart.add_data(valueOxygen, titles_from_data=True)
                oxygenChart.set_categories(valueTime)
                # stylesheet
                oxygenChart.x_axis.title = "Time(s)"
                oxygenChart.y_axis.title = "O2(%)"
                oxygenChart.height = 13
                oxygenChart.width = 19
                
                ws.add_chart(oxygenChart, "B6")

                carbonDioxideChart = LineChart()
                carbonDioxideChart.add_data(valueCarbonDioxide, titles_from_data=True)
                carbonDioxideChart.set_categories(valueTime)
                # stylesheet
                carbonDioxideChart.x_axis.title = "Time(s)"
                carbonDioxideChart.y_axis.title = "CO2(ppm)"
                carbonDioxideChart.height = 13
                carbonDioxideChart.width = 19

                ws.add_chart(carbonDioxideChart, "M6")

                heliumChart = LineChart()
                heliumChart.add_data(valueHelium, titles_from_data=True)
                heliumChart.set_categories(valueTime)
                # stylesheet
                heliumChart.x_axis.title = "Time(s)"
                heliumChart.y_axis.title = "He(%)"
                heliumChart.height = 13
                heliumChart.width = 19

                ws.add_chart(heliumChart, "X6")
            except:
                pass

        elif csvDataSheet.type == "total":
            ws = wb.active
            alignNum = 4 # ID칸 시작이 D열이다 (4씩 추가된다)

            valueTime = Reference(ws, min_row=2, max_row=ws.max_row, min_col=3, max_col=3)

            row = 2
            col = 4
            for _ in range(10):
                valueOxygen = Reference(ws, min_row=1, max_row=ws.max_row, min_col=col + 1, max_col=col + 1)
                valueCarbonDioxide = Reference(ws, min_row=1, max_row=ws.max_row, min_col=col + 2, max_col=col + 2)
                valueHelium = Reference(ws, min_row=1, max_row=ws.max_row, min_col=alignNum + 3, max_col=alignNum + 3)

                ws.column_dimensions[get_column_letter(alignNum)].width = 65

                oxygenChart = LineChart()
                oxygenChart.add_data(valueOxygen, titles_from_data=True)
                oxygenChart.set_categories(valueTime)
                # stylesheet
                oxygenChart.x_axis.title = "Time(s)"
                oxygenChart.y_axis.title = "O2(%)"
                oxygenChart.height = 13
                oxygenChart.width = 19
                
                ws.add_chart(oxygenChart, f"{get_column_letter(alignNum)}{row}")

                carbonDioxideChart = LineChart()
                carbonDioxideChart.add_data(valueCarbonDioxide, titles_from_data=True)
                carbonDioxideChart.set_categories(valueTime)
                # stylesheet
                carbonDioxideChart.x_axis.title = "Time(s)"
                carbonDioxideChart.y_axis.title = "CO2(ppm)"
                carbonDioxideChart.height = 13
                carbonDioxideChart.width = 19

                ws.add_chart(carbonDioxideChart, f"{get_column_letter(alignNum)}{row + 22}")

                heliumChart = LineChart()
                heliumChart.add_data(valueHelium, titles_from_data=True)
                heliumChart.set_categories(valueTime)
                # stylesheet
                heliumChart.x_axis.title = "Time(s)"
                heliumChart.y_axis.title = "He(%)"
                heliumChart.height = 13
                heliumChart.width = 19

                ws.add_chart(heliumChart, f"{get_column_letter(alignNum)}{row + 44}")
                
                col += 4
                # 5*6 배열 사용시 주석해제
                if alignNum >= 20:
                    alignNum = 4
                    row += 68
                    continue
                alignNum += 4
                # 5*6 여기까지
                

        elif csvDataSheet.type == "CO2" or csvDataSheet.type == "He" or csvDataSheet.type == "O2":
            ws = wb.active
            value = Reference(ws, min_row=1, max_row=ws.max_row, min_col=6, max_col=6)
            valueTime = Reference(ws, min_row=2, max_row=ws.max_row, min_col=3, max_col=3)

            chart = LineChart()
            chart.add_data(value, titles_from_data=True)
            chart.set_categories(valueTime)
            # stylesheet
            type_kor_name = {'CO2':'이산화탄소(ppm)', 'He':'헬륨(%)', 'O2':'산소(%)'}
            chart.title = type_kor_name[csvDataSheet.type]
            chart.x_axis.title = "Time(s)"
            # chart.y_axis.title = '이산화탄소(ppm)' if csvDataSheet.type == 'CO2' else '헬륨(%)'
            chart.height = 13
            chart.width = 19
            ws.add_chart(chart, "B6")


def chart_process_for_meter_wb(wb: Workbook, chart):
    ws = wb.active
    row_count = 4
    while True:
        if ws[f'B{row_count}'].value == '' or ws[f'B{row_count}'].value == None:
            ws.add_chart(chart, f'D{row_count}')
            break
        else:
            row_count += 22
    # return wb


if __name__ == "__main__":
    pass